# Overview of version: This generates the image summary INDIVIDUALLY based on text extracted from CSV FILE ONLY and output by matching to existing excel file INDIVIDUALLY.
from google.oauth2 import service_account
from googleapiclient.discovery import build
import io
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from google.oauth2 import service_account
from PIL import Image
import pytesseract
import google.generativeai as genai
import os
import pandas as pd
import nltk
from nltk.corpus import words

# Authenticate and build the google Drive API client
'''
Requires user specific inputs.

Replace SERVICE_ACCOUNT_FILE with the path to your JSON key file.
'''
SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = ''
creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
service = build('drive', 'v3', credentials=creds)

# Authenticate and build the gemini API client
'''
Requires user specific inputs.

Replace GEMINI_API_KEY with your gemini api key obtained.
'''
GEMINI_API_KEY = ''
genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel('gemini-1.5-flash')

# Function to list folders in shared drive
def list_folders_in_drive(service):
    """
    Lists all folders in the Google Drive shared with the provided service.
    
    Args:
    service: Authorized Google Drive API service instance.

    Returns:
    list: A list of dictionaries containing folder id and name.
    """
    results = service.files().list(q="mimeType = 'application/vnd.google-apps.folder'",
                                   pageSize=100, fields="files(id, name)").execute()
    return results.get('files', [])

# Function to find the fileid GIVEN filename
def find_file_id(file_name, folder_id=None):
    """
    Finds the file ID of a file by its name in the specified folder (optional).
    
    Args:
    file_name (str): The name of the file to find.
    folder_id (str, optional): The ID of the folder to search within. Defaults to None.

    Returns:
    str: The file ID of the found file, or None if not found.
    """
    # find using exact file_name provided
    query = f"name='{file_name}'"
    # if folderid provided (optional), include it in the search criteria
    if folder_id:
        query += f" and '{folder_id}' in parents"
    
    # variable to contain search results from google api
    results = service.files().list(q=query, fields="files(id, name)").execute()
    # extracts the list of files from result
    items = results.get('files', [])

    # if no files found
    if not items:
        print(f'File {file_name} not found.')
        return None

    # Assumes that there's only one file with this name
    file_id = items[0]['id']
    print(f'File {file_name} found with ID: {file_id}')
    return file_id

# Function to upload excel to drive
def upload_excel_to_drive(service, folder_id, file_name):
    """
    Uploads an Excel file to Google Drive in the specified folder.
    
    Args:
    service: Authorized Google Drive API service instance.
    folder_id (str): The ID of the folder to upload to.
    file_name (str): The name of the file to upload.
    
    Returns:
    None
    """
    file_metadata = {
        'name': file_name,
        'parents': [folder_id],
        'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    media = MediaFileUpload(file_name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    service.files().create(body=file_metadata, media_body=media, fields='id').execute()

# Function to match and merge the dataframes
def merge_excel_files(structured_df, excel_summaries_df):
    """
    Merges two dataframes by matching file names from an Excel summary dataframe with a structured dataframe.
    
    Args:
    structured_df (DataFrame): The dataframe containing structured data.
    excel_summaries_df (DataFrame): The dataframe containing Excel summaries to merge.
    
    Returns:
    DataFrame: The updated structured dataframe with merged summaries.
    """
    # Iterate through each row in the excel summaries dataframe
    for _, summary_row in excel_summaries_df.iterrows():
        file_name = summary_row['File Name']
        summary = summary_row['Summary']
        
        # Iterate through the structured dataframe to find a match
        for idx, structured_row in structured_df.iterrows():
            files = structured_row['Files'].split(', ')
            if any(file_name in file for file in files):
                # If a match is found, append the summary to the 'Summaries' column
                if pd.isna(structured_df.at[idx, 'Summaries']):
                    structured_df.at[idx, 'Summaries'] = summary
                else:
                    structured_df.at[idx, 'Summaries'] += "\n" + summary
    
    return structured_df

# Function to match and merge the dataframes SECOND TIME (for image summaries)
def second_merge_excel_files(combined_df, image_summaries_df):
    """
    Merges two dataframes by matching image names from an image summaries dataframe with a combined dataframe.
    
    Args:
    combined_df (DataFrame): The dataframe containing combined data.
    image_summaries_df (DataFrame): The dataframe containing image summaries to merge.
    
    Returns:
    DataFrame: The updated combined dataframe with merged image summaries.
    """
    # Iterate through each row in the image summaries dataframe
    for _, summary_row in image_summaries_df.iterrows():
        image_name = summary_row['Image Name']
        summary = summary_row['Summary']
        
        # Iterate through the structured dataframe to find a match
        for idx, combined_row in combined_df.iterrows():
            # Check if 'Summaries' column is empty or has the error message
            if pd.isna(combined_row['Summaries']) or combined_row['Summaries'] == "Error generating summary.":
                files = combined_row['Files'].split(', ')
                if any(image_name in file for file in files):
                    # If a match is found, append the summary to the 'Summaries' column
                    if pd.isna(combined_df.at[idx, 'Summaries']):
                        combined_df.at[idx, 'Summaries'] = summary
                    else:
                        combined_df.at[idx, 'Summaries'] += "\n" + summary
    
    return combined_df


# Function to delete xlsx file in drive
def delete_existing_xlsx_files(service, folder_id):
    """
    Deletes existing .xlsx files in the specified folder on Google Drive.
    
    Args:
    service: Authorized Google Drive API service instance.
    folder_id (str): The ID of the folder to search and delete files from.
    
    Returns:
    None
    """
    # Query to find existing .xlsx files in the specified folder
    query = f"'{folder_id}' in parents and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'"
    
    # Fetch the list of files matching the query
    files = service.files().list(q=query, pageSize=10, fields="files(id, name)").execute().get('files', [])
    
    # Check if there are files to delete
    if not files:
        print(f"No .xlsx files found in folder {folder_id}")
        return
    
    # Iterate over the list of files and attempt to delete each
    for file in files:
        try:
            service.files().delete(fileId=file['id']).execute()
            print(f"Deleted existing file {file['name']} from folder {folder_id}")
        except Exception as e:
            print(f"Failed to delete file {file['name']} from folder {folder_id}: {e}")

# Function to delete file by name
def delete_file_by_name(file_name, folder_id=None):
    """
    Deletes a file from Google Drive by its name and optional folder ID.
    
    Args:
    file_name (str): The name of the file to delete.
    folder_id (str, optional): The ID of the folder to search in. Defaults to None.
    
    Returns:
    None
    """
    file_id = find_file_id(file_name, folder_id)
    if file_id:
        try:
            service.files().delete(fileId=file_id).execute()
            print(f'Successfully deleted file with name: {file_name}')
        except Exception as e:
            print(f'Error deleting file {file_name}: {e}')

# Function to download file from drive and save locally (efficient handling of large files by downloading them in chunks)
def download_json_file(service, file_id, file_name):
    """
    Downloads a JSON file from Google Drive and saves it locally.
    
    Args:
    service: Authorized Google Drive API service instance.
    file_id (str): The ID of the file to download.
    file_name (str): The local file name to save the downloaded file as.
    
    Returns:
    None
    """
    request = service.files().get_media(fileId=file_id)
    fh = io.FileIO(file_name, 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()

# Function to process a JSON file: extracts paths, objects, and texts, and generates an Excel file containing this information.
def process_json_and_generate_excel(json_file, excel_file):
    """
    Processes a JSON file to extract paths, objects, and texts, and generates an Excel file with this data.
    
    Args:
    json_file (str): The path to the input JSON file.
    excel_file (str): The path to save the output Excel file.
    
    Returns:
    None
    """
    with open(json_file, 'r') as f:
        data = json.load(f)

    # eg filePaths:'Path'
    general_paths = {}
    # eg filePaths:[]
    specific_paths = {}
    #eg filePaths:[]
    objects_text = {}
    # eg filePaths:[]
    texts = {}
    # eg filePaths:[]
    object_file = {}

    for element in data['elements']:
        if 'filePaths' in element:
            key = tuple(element['filePaths'])
            value = element['Path']
            general_paths[key] = value + '/'
            object_file[key] = element['ObjectID']

    for key in general_paths.keys():
        objects_text[key] = []
        texts[key] = []
        specific_paths[key] = []

    for key, value in general_paths.items():
        for element in data['elements']:
            if (element['Path'] + '/' == value or 
                    (element['Path'].startswith(value) and len(element['Path']) > len(value))):
                if 'Text' in element:
                    texts[key].append(element['Text'])
                    objects_text[key].append(element['ObjectID'])
                    specific_paths[key].append(element['Path'])

    combined_dict = {}
    for key in general_paths.keys():
        combined_dict[key] = (object_file[key], general_paths[key], specific_paths[key], texts[key], objects_text[key])

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['Files', 'ObjectID_file', 'General Paths', 'Specific Paths', 'Texts', 'ObjectIDs_text', 'Summaries'])

    for key, value in combined_dict.items():
        # summary = summarize_content(value[2], value[3], value[1])
        ws.append([', '.join(key)] + [str(val) for val in value])

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)
            max_length = max(max_length, len(str(cell.value) or ""))

        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    wb.save(excel_file)

nltk.download('words')

# Set of valid English words
english_vocab = set(words.words())

# Function to detect gibberish
# Adjust threshold as needed (now it's 90%)
def is_gibberish(text, threshold=0.9):
    """
    Detects whether the given text is gibberish by comparing the number of valid English words 
    to the total words. A threshold can be set to determine the proportion of valid words.

    Args:
        text (str): The text to analyze for gibberish content.
        threshold (float): The threshold for valid words proportion (default is 0.9).

    Returns:
        bool: True if the text is considered gibberish, False otherwise.
    """
    # Split the text into words and filter out short words
    words_in_text = [word for word in text.split() if len(word) > 2]

    # Check if there's a significant presence of valid English words
    valid_words = [word for word in words_in_text if word.lower() in english_vocab]
    
    # Determine gibberish based on the stricter threshold
    return len(valid_words) < len(words_in_text) * threshold

# Function to download image from Google Drive
def download_image_file(service, file_id, file_name):
    """
    Downloads an image file from Google Drive using the provided file ID and saves it to the 
    specified file name.

    Args:
        service (googleapiclient.discovery.Resource): The authenticated Google Drive service instance.
        file_id (str): The ID of the file to download from Google Drive.
        file_name (str): The local name to save the file as.

    Returns:
        str: The local path where the file was saved.
    """
    request = service.files().get_media(fileId=file_id)
    fh = io.FileIO(file_name, 'wb')
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    return file_name

# Function to extract text from image using tesseract
def extract_text_from_image(image_file):
    """
    Extracts text from an image using Tesseract OCR. Attempts to rotate the image if the extracted 
    text is gibberish, trying up to 3 rotations (90, 180, and 270 degrees).

    Args:
        image_file (str): The path to the image file from which to extract text.

    Returns:
        str: The extracted text from the image.
    """
    img = Image.open(image_file)
    text = pytesseract.image_to_string(img)

    # If the text is gibberish, try rotating the image and re-extracting
    for i in range(3):  # Try up to 3 rotations (90, 180, 270 degrees)
        if not is_gibberish(text):
            break  # Exit loop if valid text is found

        img = img.rotate(90, expand=True)  # Rotate the image 90 degrees clockwise
        text = pytesseract.image_to_string(img)

    print(text)
    return text

# Function to generate summary using gemini given text
def generate_image_summary(extracted_text):
    """
    Summarizes the extracted text from images (flow charts, tables, questionnaires, or figures). 
    The summary follows specific formats depending on the type of image content.

    Args:
        extracted_text (str): The extracted text from the image.

    Returns:
        str: A concise summary of the extracted content.
    """
    prompt = (
        "The following text is extracted from either images of flow charts, tables, questionnaires, or figures."
        "Firstly, determine and remember what type of image the text was extracted from."
        "Next, I would like you to summerise the text extracted. Ensure that the summary reflects the image structure and all textual information extracted"
        "Organize the summary as follows:\n"
        "1. For flow charts: Describe the flow and connections between boxes. Summarize the steps and decisions.\n"
        "2. For tables: Determine the column and row structure. Then, for each row of the table, explain the text in the row with relevance to the columns. Include the row contents succinctly.\n"
        "3. For questionnaires: Summarize the questions and the options for the questions. Ensure clarity and coherence.\n\n"
        "4, For figures, briefly describe the texts extracted from the image."
        "Ignore any gibberish or non-sensical content. Do not include any extraneous information or explanations. Provide a complete, clear, and concise summary of the extracted content.\n\n"
        f"Text: {extracted_text}"
    )

    try:
        response = model.generate_content(prompt)
        
        # Debugging: Print the raw response
        print("API Response:", response)
        
        # Safely access response content
        if hasattr(response, 'text') and response.text:
            return response.text.strip()
        elif 'content' in response and response['content']:
            return response['content'].strip()
        else:
            return "Summary could not be generated."
    except Exception as e:
        print(f"Error generating summary: {e}")
        return "Error generating summary."
    
# Function to recursively process images in all folders and subfolders and output as a temp excel
def process_images(service, folder_id, worksheet):
    """
    Recursively processes all images in the specified folder and its subfolders, extracts text, 
    generates summaries, and appends the results to an Excel worksheet.

    Args:
        service (googleapiclient.discovery.Resource): The authenticated Google Drive service instance.
        folder_id (str): The ID of the Google Drive folder to process.
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to append the results to.

    Returns:
        None
    """
    query = f"'{folder_id}' in parents and mimeType = 'application/vnd.google-apps.folder'"
    subfolders = service.files().list(q=query, pageSize=100, fields="files(id, name)").execute().get('files', [])
    
    for subfolder in subfolders:
        process_images(service, subfolder['id'], worksheet)
    
    query = f"'{folder_id}' in parents and mimeType contains 'image/'"
    images = service.files().list(q=query, pageSize=100, fields="files(id, name)").execute().get('files', [])
    
    for image in images:
        image_name = image['name']
        print(f"Processing {image_name}...")
        local_image_file = download_image_file(service, image['id'], image_name)
        extracted_text = extract_text_from_image(local_image_file)
        summary = generate_image_summary(extracted_text)
        
        # Append to Excel worksheet
        worksheet.append([image_name, summary])
        os.remove(local_image_file)

# Function to download an .xlsx file from Google Drive
def download_xlsx_file(service, file_id, file_name):
    """
    Downloads an Excel file from Google Drive using the provided file ID and saves it locally.

    Args:
        service (googleapiclient.discovery.Resource): The authenticated Google Drive service instance.
        file_id (str): The ID of the Excel file to download from Google Drive.
        file_name (str): The local name to save the file as.

    Returns:
        str: The local path where the file was saved.
    """
    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
        print(f"Download {int(status.progress() * 100)}%.")
    fh.seek(0)
    with open(file_name, 'wb') as f:
        f.write(fh.read())
    return file_name


def extract_text_from_xlsx(file_path):
    """
    Reads an Excel file and returns its contents. It skips files that do not start with 'file' in their name.

    Args:
        file_path (str): The path to the Excel file to extract data from.

    Returns:
        pandas.DataFrame: The data from the Excel file as a DataFrame, or None if the file is invalid.
    """
    # Get the file name from the file path
    file_name = os.path.basename(file_path)
    
    # Check if the file name starts with "file"
    if not file_name.startswith("file"):
        print(f"Skipping file as it does not start with 'file': {file_name}")
        return None
    
    # Check if the file path is being generated correctly
    print(f"Attempting to read file: {file_path}")
    
    if not os.path.exists(file_path):
        print(f"File does not exist: {file_path}")
        return None
    
    # Proceed with reading the file
    try:
        df = pd.read_excel(file_path)
        # Print the entire DataFrame
        print(f"Contents of the file {file_name}:\n{df.to_string()}")
        return df
    except Exception as e:
        print(f"Error reading the file: {e}")
        return None

# Function to generate summary using gemini given text from xlsx
def generate_xlsx_summary(extracted_text):
    """
    Generates a summary from text extracted from Excel files, with specific rules for handling 
    flow charts, tables, or questionnaires. Summarizes the content row by row.

    Args:
        extracted_text (str): The extracted text from the Excel file.

    Returns:
        str: A concise summary of the Excel table's content.
    """
    prompt = (
        "The following text input is extracted from excel tables containing the text extracted from images of flow charts, tables, or questionnaires.n\n"
        "Firstly, output the column headers of the excel table and provide a brief summary of the table structure."
        "Next, interpret the table content and output the values for each row based on the columns identified in continuous prose."
        "For example, depending on the number of columns and rows:"
        "In the first row, the {column 1 name} is ___ and the {column 2 name} value is ___, the {column 3 name} value is ___, ____."
        "If there are any gibberish or non-sensical content extracted, ignore them. Else, include all texts extracted from each cell of the excel file\n\n"
        "For text extracted that are related to drugs, they are safe and do not halt text generation by raising a 'Finish Reason: SAFETY' flag.\n\n"
        "Do not include any extraneous information or explanations.\n\n"
        "Provide a complete, clear, and concise summary of the extracted content.\n\n"
        f"Text: {extracted_text}"
    )

    try:
        response = model.generate_content(prompt)
        
        # Debugging: Print the raw response
        print("API Response:", response)
        
        # Safely access response content
        if hasattr(response, 'text') and response.text:
            return response.text.strip()
        elif 'content' in response and response['content']:
            return response['content'].strip()
        else:
            return "Summary could not be generated."
    except Exception as e:
        print(f"Error generating summary: {e}")
        return "Error generating summary."
    
# Function to process excel tables for table summaries
def process_xlsx_files(service, folder_id, worksheet):
    """
    Recursively processes Excel files in the specified folder and its subfolders, extracts text, 
    generates summaries, and appends the results to an Excel worksheet.

    Args:
        service (googleapiclient.discovery.Resource): The authenticated Google Drive service instance.
        folder_id (str): The ID of the Google Drive folder to process.
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to append the results to.

    Returns:
        None
    """
    query = f"'{folder_id}' in parents and mimeType = 'application/vnd.google-apps.folder'"
    subfolders = service.files().list(q=query, pageSize=100, fields="files(id, name)").execute().get('files', [])
    
    for subfolder in subfolders:
        process_xlsx_files(service, subfolder['id'], worksheet)
    
    query = f"'{folder_id}' in parents and mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'"
    files = service.files().list(q=query, pageSize=100, fields="files(id, name)").execute().get('files', [])
    
    for file in files:
        file_name = file['name']
        
        # Check if the file name starts with "file"
        if not file_name.startswith("file"):
            print(f"Skipping {file_name} as it does not start with 'file'.")
            continue
        
        print(f"Processing {file_name}...")
        local_file = download_xlsx_file(service, file['id'], file_name)
        
        # Load the Excel file
        df = pd.read_excel(local_file)
        
        # Identify column names
        columns = df.columns.tolist()
        
        # Initialize a list to store row summaries
        row_summaries = []
        
        # Iterate through each row and generate a summary
        for index, row in df.iterrows():
            # Create a summary dynamically based on column names
            summary_parts = [f"In row {index + 1},"]
            for i, col in enumerate(columns):
                # Strip leading/trailing whitespace and remove newline characters
                value = str(row[col]).strip().replace('\n', ' ').replace('\r', '')
                if value:
                    # Append each column-value pair without extra line breaks
                    summary_parts.append(f"The {col} is: {value}")
                    # Add a period after each pair except the last one
                    if i < len(columns) - 1:
                        summary_parts.append(". ")
                    else:
                        summary_parts.append(".")
            # Join all parts into a single summary string
            summary = " ".join(summary_parts).strip()
            row_summaries.append(summary)

        # Combine all row summaries into a single summary string
        full_summary = "\n".join(row_summaries)
        
        # Print the summary to check
        print(f"Summary for {file_name}:\n{full_summary}\n")
        
        # Append to Excel worksheet
        worksheet.append([file_name, full_summary])
        
        # Clean up downloaded file after processing
        os.remove(local_file)


# Main application
def main():
    '''
    Requires user specific inputs.

    The google drive folder of interest in this script is the "PDF_Extrated data" subfolder under the "Data" folder.

    Replace root_folder_id with the folder id of the google drive folder of interest.
    '''

    # Replace with the folderid of the folder you are interacting with
    root_folder_id = '1r38pL-SjbkwYBoK5EF1_Ou4sxb3iw0H7'

    # Get the IDs for the folders within the root folder
    query = f"'{root_folder_id}' in parents and mimeType = 'application/vnd.google-apps.folder'"
    main_folders = service.files().list(q=query, pageSize=100, fields="files(id, name)").execute().get('files', [])
    print("Main Folders:", main_folders)

    for folder in main_folders:
        # Check if folder contains an Excel file starting with "Combined"
        query_combined = f"'{folder['id']}' in parents and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'"
        xlsx_files = service.files().list(q=query_combined, pageSize=10, fields="files(id, name)").execute().get('files', [])
        
        # If any file starts with "Combined", skip the loop
        if any(file['name'].startswith('Combined') for file in xlsx_files):
            print(f"Skipping folder {folder['name']} as it contains a file starting with 'Combined'")
            continue

        # Process new .json files and generate .xlsx file
        query = f"'{folder['id']}' in parents and mimeType='application/json'"
        files = service.files().list(q=query, pageSize=10, fields="files(id, name)").execute().get('files', [])

        for file in files:
            json_file = file['name']

            if json_file != 'structuredData.json':
                continue

            download_json_file(service, file['id'], json_file)
            
            excel_file = json_file.replace('.json', '.xlsx')
            process_json_and_generate_excel(json_file, excel_file)
            
            upload_excel_to_drive(service, folder['id'], excel_file)
            print(f"Processed and uploaded {excel_file} to folder {folder['name']}")
            structured_df = pd.read_excel(excel_file)
        
        # Create a new Excel workbook and worksheet for excel table summaries
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "xlsx Summaries"
        worksheet.append(["File Name", "Summary"])

        # Process xlsx in the current folder
        process_xlsx_files(service, folder['id'], worksheet)

        # Save the xlsx_summary file with the folder's name
        xlsx_summary_file_name = f"xlsx_Summaries_{folder['name']}.xlsx"
        workbook.save(xlsx_summary_file_name)
        print(f"xlsx summaries saved to {xlsx_summary_file_name}")

        # Upload the xlsx_summary file to the Google Drive subfolder where the images were located
        upload_excel_to_drive(service, folder['id'], xlsx_summary_file_name)

        # Combine Excel files
        xlsx_summaries_df = pd.read_excel(xlsx_summary_file_name)
        merged_df = merge_excel_files(structured_df, xlsx_summaries_df)
        merged_excel_file_name = f"Combined_{folder['name']}.xlsx"
        merged_df.to_excel(merged_excel_file_name, index=False)

        # Create a new Excel workbook and worksheet for image summaries
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Image Summaries"
        worksheet.append(["Image Name", "Summary"])

        # Process images in the current folder
        process_images(service, folder['id'], worksheet)

        # Save the Excel file with the folder's name
        image_summary_file_name = f"Image_Summaries_{folder['name']}.xlsx"
        workbook.save(image_summary_file_name)

        # Upload the Excel file to the Google Drive subfolder where the images were located
        upload_excel_to_drive(service, folder['id'], image_summary_file_name)

        combined_df = pd.read_excel(merged_excel_file_name)
        image_summaries_df = pd.read_excel(image_summary_file_name)
        merged_df = second_merge_excel_files(combined_df,image_summaries_df)
        merged_excel_file_name = f"Combined_{folder['name']}.xlsx"
        merged_df.to_excel(merged_excel_file_name, index=False)
        upload_excel_to_drive(service, folder['id'], merged_excel_file_name)

        # Delete the individual files
        delete_file_by_name(excel_file)
        delete_file_by_name(xlsx_summary_file_name)
        delete_file_by_name(image_summary_file_name)

if __name__ == '__main__':
    main()